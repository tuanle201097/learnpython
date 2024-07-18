import os
import openpyxl

def getInfor(file_path,input_num,cases,index,colums):
    # Load workbook (bảng tính)
    wb = openpyxl.load_workbook(file_path)
    # Chọn sheet cụ thể (nếu cần)
    sheet = wb['Sheet1']  # Thay 'Sheet1' bằng tên của sheet bạn muốn đọc
    if cases == 1:
        # Đọc dữ liệu từng ô
        # Ví dụ: Đọc giá trị từ ô c2
        cellurl_position = f'C{input_num}'  # Tạo vị trí ô dạng 'C{input_num}'
        cellname_position = f'E{input_num}'  # Tạo vị trí ô dạng 'E{input_num}'
        print(f"Vị trí ô: {cellurl_position}")
        print(f"Vị trí name: {cellname_position}")
        urlvalue = sheet[cellurl_position].value
        namevalue = sheet[cellname_position].value
        # Đóng workbook sau khi sử dụng
        wb.close()
    elif cases == 2:
        urlvalue = sheet.cell(row=index, column=colums).value  # Lấy giá trị ô trong cột F
        namevalue = None
    else:
        urlvalue = sheet.cell(row=index, column=colums).value  # Lấy giá trị ô trong cột F
        namevalue = sheet.cell(row=index+1, column=colums).value  # Lấy giá trị ô trong cột F
    return urlvalue, namevalue  